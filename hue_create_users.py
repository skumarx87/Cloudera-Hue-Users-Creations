from openpyxl import load_workbook
import time
import ssl
import kerberos
import requests
import json
import sys

from bs4 import BeautifulSoup


class KerberosTicket:
    def __init__(self, service):
        __, krb_context = kerberos.authGSSClientInit(service)
        kerberos.authGSSClientStep(krb_context, "")
        self._krb_context = krb_context
        self.auth_header = ("Negotiate " +kerberos.authGSSClientResponse(krb_context))
    def verify_response(self, auth_header):
        # Handle comma-separated lists of authentication fields
        for field in auth_header.split(","):
            kind, __, details = field.strip().partition(" ")
            print(kind)
            if kind.lower() == "negotiate":
                auth_details = details.strip()
                break
        else:
            raise ValueError("Negotiate not found in %s" % auth_header)
        # Finish the Kerberos handshake
        krb_context = self._krb_context
        if krb_context is None:
            raise RuntimeError("Ticket already used for verification")
        self._krb_context = None
        kerberos.authGSSClientStep(krb_context, auth_details)
        kerberos.authGSSClientClean(krb_context)


class readxls:
	def __init__(self,hue_server,port):
		self.hue_host='https://'+hue_server+':'+port
		print("HTTP/"+hue_server+"@TANU.COM")
		self.krb = KerberosTicket("HTTP/"+hue_server+"@TANU.COM")
		self.headers = {"Authorization": self.krb.auth_header,'referer': self.hue_host+'/home'}
		self.session = requests.Session()
		self.r = self.session.get(self.hue_host+"/home", headers=self.headers,verify="/var/ssl/truststore.pem")
		self.session.headers.update({'referer': self.hue_host+'/home'})
		self.cooki=self.session.cookies['csrftoken']
		self.banner_message("INFO: This is script Tested on only Hue 4.0")
	def banner_message(self,message):
		print("------------------------------------------------------------------")
		print(message)
		print("------------------------------------------------------------------")

	def create_group(self):
		self.banner_message("Creating Hue groups")
		for group in self.HUE_GROUPS:
			print("Creating Hue group "+str(group))
        		payload={'csrfmiddlewaretoken':self.cooki,'name':group}
        		self.session.post(self.hue_host+'/useradmin/groups/new', headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem",data=payload)

	def map_user_group(self):
		self.banner_message("Mapping Hue groups")
                for user,groups in self.USER_MAP_LIST.iteritems():
			r=self.session.get(self.hue_host+"/useradmin/users/edit/"+user, headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem")		
			if r.status_code != 200:
				print("ERROR !!!!!! skipping user {}".format(user))
				continue
			soup = BeautifulSoup(r.text, "html.parser")
			hue_groups_id = {}
			for option in soup.find_all('option'):
				hue_groups_id[str(option.text).strip().encode('utf-8')] = str(option['value']).strip().encode('utf-8')
			first_name=soup.find('input',{'id':'id_first_name'}).get('value')
			last_name=soup.find('input',{'id':'id_last_name'}).get('value')
                	email=soup.find('input',{'id':'id_email'}).get('value')
			self.cooki=self.session.cookies['csrftoken']
			if isinstance (groups,str):
				payload={'csrfmiddlewaretoken':self.cooki,'username':user,'groups':hue_groups_id[groups],'is_active':'on','first_name':str(first_name),'last_name':str(last_name),'email':str(email)}
			if isinstance(groups,list):
				arry=[]
				for group in groups:
					arry.append(hue_groups_id[group])
				payload={'csrfmiddlewaretoken':self.cooki,'username':user,'groups':arry,'is_active':'on','first_name':str(first_name),'last_name':str(last_name),'email':str(email)}
					
			print("Mapping user {} into group(s) {}".format(user,groups))
			c=self.session.post(self.hue_host+"/useradmin/users/edit/"+user, headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem",data=payload)
	def parse_xls(self):
		UNIX_GROUP_COL_ID='E'
		NT_ID_COL='A'
		STARTING_RECORD_NO=5
		wb2 = load_workbook('/cs/cmluat/hue_user_creation/Sentry_list.xlsx')
		ws = wb2.get_sheet_by_name(name = 'Sentry_Role_mapping')
		print("calling parse xls")	
		GROUPS=[];
		USER_MAP={};
		while True:
			uid=ws[NT_ID_COL+str(STARTING_RECORD_NO)].value
			ugroup=ws[UNIX_GROUP_COL_ID+str(STARTING_RECORD_NO)].value
			next_uid=ws[NT_ID_COL+str(STARTING_RECORD_NO+1)].value
			next_group=ws[UNIX_GROUP_COL_ID+str(STARTING_RECORD_NO+1)].value
                        if ugroup is None:
                                break
			if next_group is None:
				USER_MAP[uid.strip().encode('utf-8').lower()]= ugroup.strip().encode('utf-8').lower();
			if next_uid is not None:
				USER_MAP[uid.strip().encode('utf-8').lower()]= ugroup.strip().encode('utf-8'.lower());
			if next_uid is None:
				USER_GROUPS=[]
				USER_GROUPS.append(str(ugroup).strip().encode('utf-8').lower())
				start_record=STARTING_RECORD_NO+1
				while True:
					ugroup=ws[UNIX_GROUP_COL_ID+str(start_record)].value
					newuid=ws[NT_ID_COL+str(start_record)].value
					if newuid is None:	
						USER_GROUPS.append(str(ugroup).strip().encode('utf-8').lower())
					if ugroup is None:
						break
					if newuid is not None:
						
						USER_MAP[uid.strip().encode('utf-8').lower()]=USER_GROUPS
						STARTING_RECORD_NO = start_record - 1;
						break
					start_record +=1 
			if ugroup is not None:
				GROUPS.append(ugroup.strip().encode('utf-8').lower())
			STARTING_RECORD_NO += 1	
		self.HUE_GROUPS=list(set(GROUPS)) ##Remove duplicat items
		self.USER_MAP_LIST=USER_MAP
	'''
        def create_hue_users(self):
		#print("Calling Create hue users ")
		self.banner_message("Calling Create hue users")
		for user,groups in self.USER_MAP_LIST.iteritems():
			print("Adding user {} into Hue".format(user))
			payload={'csrfmiddlewaretoken':self.cooki,'username_pattern':user,'ensure_home_directory':'on'}
			self.session.post(self.hue_host+'/useradmin/users/add_ldap_users', headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem",data=payload)
			
	'''	
        def create_hue_users(self,user):
                #print("Calling Create hue users ")
                self.banner_message("Calling Create hue user")
		payload={'csrfmiddlewaretoken':self.cooki,'username_pattern':user,'ensure_home_directory':'on'}
		self.session.post(self.hue_host+'/useradmin/users/add_ldap_users', headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem",data=payload)
                print("Adding user {} into Hue".format(user))

	def sync_users(self):
		self.banner_message("Running sync users")
		#print("Running sync users")
		r=self.session.get(self.hue_host+"/useradmin/users", headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem")
		soup = BeautifulSoup(r.text, "html.parser")
		data = []
		table = soup.find('table', attrs={'class':'datatables'})		
		table_body = table.find('tbody')
		rows = table_body.find_all('tr')
		for row in rows:
			cols = row.find_all('td')
			for col in cols:
				c=col.find('div')
				if c is not None:
					ss=c['data-id'].strip().encode('utf-8').lower()
			cols = [ele.text.strip().encode('utf-8') for ele in cols]
			cols.append(ss)
			#print(cols)
			data.append([ele for ele in cols if ele])
		for u in data:
			user=str((u[0]).strip().encode('utf-8'))
			id=u[-1]
			if user in self.USER_MAP_LIST:
				print("User exist")
			else:
				print("Deleting user {} with data id {}".format(user,id))
				payload={'csrfmiddlewaretoken':self.cooki,'user_ids':id}
				self.session.post(self.hue_host+'/useradmin/users/delete', headers=self.session.headers,cookies=self.session.cookies,verify="/var/ssl/truststore.pem",data=payload)
				#print("user {} not exist".format(user))
		for user,groups in self.USER_MAP_LIST.iteritems():
			check_exist=[hue_user for hue_user in data if user in hue_user]
			if len(check_exist) >= 1:
				print("User already availabe in Hue {}".format(user)) 
				print(check_exist[0])
			elif len(check_exist) == 0:
				print("user {} not exist".format(user))
				self.create_hue_users(user)
if ((len(sys.argv) != 3)):		
	print("Usage : Hue server hostname and port")
	sys.exit(0)
hostname=sys.argv[1]
port=sys.argv[2]
read = readxls(hostname,port) #Initialize python class
read.parse_xls()  #Parse the xls sheet users and group Informations.
#read.create_hue_users() #Create the users from parsed XLS sheet.
read.sync_users()     #Delete the users if users not exists in the sheet and crete new users.
read.create_group()   #Create groups in hue
read.map_user_group() #Map the repsective groups to users.
